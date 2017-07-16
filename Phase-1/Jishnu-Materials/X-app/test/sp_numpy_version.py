#import required libraries

from openpyxl import load_workbook
import numpy as np


class orange:

	def __init__(self, input_file, budget):
		#This is the input file to the webapp which contains 
	    #[1] Coeficients for each group
		#[2] Sampling cost for each group
		self.input_file = input_file 
		
		# total amount to be used to conduct experiment (experiment_budget)
		self.A = float(budget) 

	def getColumnValues(self, working_sheet, total_num_of_groups, column_index):
		columnVector = np.zeros(total_num_of_groups) # Coeficients for each group
		for i in range(0,total_num_of_groups):
			try:
				columnVector[i]=working_sheet.cell(row=i+2, column=column_index).value
				#print columnVector[i]
		 	except Exception as e: 
		 		print str(e)
		return columnVector

	def performSequentialAnalysis(self):
		#read from excel file
		workbook = load_workbook(self.input_file)
		working_sheet = workbook.get_sheet_by_name('Sheet1')
		total_groups = working_sheet.max_row-1

		print("total_groups : {}".format(total_groups))

		mu = self.getColumnValues(working_sheet, total_groups, 1)	# Mean for each group 
		sigma = self.getColumnValues(working_sheet, total_groups, 2) # Standard Deviation for each group
		C = self.getColumnValues(working_sheet, total_groups, 3) # Coeficients for each group
		a = self.getColumnValues(working_sheet, total_groups, 4) # Sampling cost for each group

		num_iterations=5
		current_iteration=1

		denominator_matrix = np.zeros((total_groups,1), float)
		test_result_matrix = np.zeros((total_groups,num_iterations),float)

		#Theoritical calulation of sample value per group
 		optimum = np.zeros((total_groups,3),float)#optimum matrix containing both optimum sample size & respective deviation

 		print C
 		print sigma
 		print sigma[0]

		for i in range(0,total_groups):
			print (self.A)*(abs(C[0][i])*sigma[0][i])
			num = (self.A)*abs(C[0][i])*sigma[0][i]
		 	for l in range(0,total_groups):
		 		denominator_matrix[l]=abs(C[0][l])*(sigma[0][l])*(np.sqrt(a[l]*a[i]))
		 	optimum[i] = num/np.sum(denominator_matrix)
		print optimum

		while current_iteration <= num_iterations:
			print("Current iteration : {}".format(current_iteration))    	
			m=2#size of pilot data for each group
			counter = 0
			pilot_data = np.zeros((1,m),float)
			sd_of_each_group = np.zeros((total_groups,1), float) #stores sd of each group when it reaches its optimum sample size
			#pilot_data generated successfully
			numerator=0
			denominator=0
			calculated_value = np.zeros((1,total_groups),float)#stores num/denom value for each group after each iteration
			state = np.zeros((1,total_groups),float)#this vector is used to determine whether that particular group achieved optimum value(i.e. 1) or not (i.e. "0")
			#generating pilot_data
			for i in range(0,total_groups):
				temp = np.random.normal(mu[i], sigma[i], m)
				#print("Generated normal rand_num : {}".format(temp))    	
				#print("Pilot_data : {}".format(pilot_data))    	
				if(pilot_data[0][0]==0):
					pilot_data = temp
				else:
					pilot_data = vstack((pilot_data,temp))

				while counter < total_groups:#this counter shows #groups that have achieved optimum value
		  			for i in range(0,total_groups):
		  				if state[0][i]==0:
		  					numerator = (self.A)*abs(C[0][i])*(np.std(pilot_data[i,]))
		   					for l in range(0,total_groups):
								if state[l]==1:
									denominator_matrix[l]=abs(C[0][l])*(sd_of_each_group[l,])*np.sqrt(a[l]*a[i])
								else:#if(state[l]==0)
									denominator_matrix[l]=abs(C[0][l])*(np.std(pilot_data[l,]))*np.sqrt(a[l]*a[i])
							calculated_value[i] = numerator/sum(denominator_matrix)
		    		###condition check
		  			for i in range(0,total_groups):
						b = calculated_value[i]
						if (m>= b and state[i]==0):
							state[i]=1
							#optimum[i,1]=m
						#optimum[i,2]=sd(pilot_data[i,])
						test_result_matrix[i,current_iteration] = m
						sd_of_each_group[i]=np.std(pilot_data[i])
						counter = counter+1
									
									
					#steps to update m & pilot_data & restart process again for non-optimized groups
		  			m=m+1 #increamenting m
		  			pilot_data = hstack((pilot_data,"0"))
		  			for i in range(0,total_groups):#updating the pilot data according to value of m
						if state[i]==0:
							pilot_data[i,m] = np.random.normal(mu[i], sigma[i], 1)
		   
		   
		   
		  	current_iteration = current_iteration + 1

  		for i in range(0,total_groups): 
  			optimum[i,2] = np.mean(test_result_matrix[i])
   			optimum[i,3] = np.std(test_result_matrix[i])
 
		print(optimum)


		 	
o = orange('input.xlsx', '50000')
o.performSequentialAnalysis()


